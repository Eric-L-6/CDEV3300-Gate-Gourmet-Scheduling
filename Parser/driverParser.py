import os
import openpyxl
from Helpers.drivers import Driver

num_skill = 19
offset = 2

class DriverParser():
    def driverInitialise(file_path: str):
        drivers = []

        # read the excel file
        workbook_path = os.path.join(os.getcwd(),"Input Data", file_path)
        wb = openpyxl.load_workbook(workbook_path)
        ws = wb['Team member list']

        skillset = []
        for i in range(3, ws.max_column + 1):
            if (ws.cell(row=1, column=i).value == "Total point"):
                break
            skillset.append(ws.cell(row=1, column=i).value)

        for i in range(2, ws.max_row + 1):
            if ws.cell(row=i, column=1).value is not None and type(ws.cell(row=i, column=1).value) != int:
                break
            if ws.cell(row=i, column=1).value is not None:
                # print(ws.cell(row=i, column=1).value)
                driver_info = []
                for j in range(1, num_skill + 3):
                    driver_info.append(ws.cell(row=i, column=j).value)
                driver = Driver(driver_info, skillset)
                drivers.append(driver)

                # print(driver)

        return drivers