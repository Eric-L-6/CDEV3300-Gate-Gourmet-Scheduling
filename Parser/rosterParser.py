from openpyxl import load_workbook
from openpyxl.styles import PatternFill
from datetime import datetime, timedelta
import os
import re

class RosterParser():
    missing_slot = "欠"
    processed_slot = "完"
    unavailable_slot = "FFFF0000"
    driver_id_map = {}
    date_index_map = {}
    dates = []
    ws = None
    current_date = None

    def __init__(self, workbook_path):
        # Load the workbook and select the active sheet
        self.workbook_path = workbook_path
        self.wb = load_workbook(self.workbook_path)
        self.ws = self.wb["30-SEP"]
        self.driver_id_map = self.initDriverIdMap()
        self.date_index_map = self.initDateIndexMap()
    
    def getMonthlyDateRange(self):
        return self.dates

    def initDriverIdMap(self):
        driver_id_map = {}
        for row in range(2, self.ws.max_row + 1):
            driver_id = self.ws.cell(row=row, column=1).value
            if driver_id is not None:
                driver_id_map[self.ws.cell(row=row, column=1).value] = row
        return driver_id_map
    
    def initDateIndexMap(self):
        date_index_map = {}
        for col in range(2, self.ws.max_column + 1):
            date = self.ws.cell(row=2, column=col).value
            # date is a datetime object
            if date is not None and type(date) == datetime:
                date_index_map[date] = col
                self.dates.append(date)
        return date_index_map
    
    def getKeyFromValue(self, d, value):
        for key, val in d.items():
            if val == value:
                return key
        return None


    def getDateRange(self):
        # get the date range from the excel file 
        return

    # Function to access a cell based on row and column
    def getCellInfo(self, row, col):

        # Get the cell based on the row and column
        cell = self.ws.cell(row=row, column=col)
        cell_info = {
            "value": cell.value,
            "fill_color": None
        }

        # Check if the cell has a fill and access its color
        if cell.fill and cell.fill.fgColor.type == "rgb":
                cell_info["fill_color"] = cell.fill.fgColor.rgb

        return cell_info
    
    def getRosterSheetList(self):
        # get the list of sheets in the excel file
        sheet_list = self.wb.sheetnames
        # Define the regex pattern for DD-MMM format
        pattern = re.compile(r'^\d{1,2}-[A-Z]{3}$')
        
        # Filter the list using the pattern
        filtered_list = [date for date in sheet_list if pattern.match(date)]
        
        return filtered_list
    
    def setSheet(self, sheet_name):
        # set the active sheet to the given sheet name
        if sheet_name not in self.wb.sheetnames:
            raise Exception(f"Sheet {sheet_name} not found in workbook")
        self.ws = self.wb[sheet_name]

    
    # write a specific format of cell into the excel file
    # all cell in the given range if not filled with red color will be filled with that specific format
    def writing(self, row, col, value=None, fill_color="00000000"):
        cell = self.ws.cell(row=row, column=col)
        
        # Write a value to the cell if provided
        # color of the value will be black
        if value is not None:
            cell.value = value
            cell.font = cell.font.copy(color="FF000000")
        
        # Change the fill color of the cell
        cell.fill = PatternFill(start_color=fill_color, end_color=fill_color, fill_type="solid")
        print(f"Writing to cell ({row}, {col}): value={value}, fill_color={fill_color}")
    
        return cell

    # Precondition: 
    # - date is a datetime object
    # - drivers is a list of driver class objects
    # Postcondition:
    # - Returns a list of available drivers based on the date
    def getAvailableDrivers(self, date:datetime, drivers:list):
        # Get the column index based on the date
        date_col = None
        for col in range(1, self.ws.max_column + 1):
            if self.ws.cell(row=2, column=col).value == date:
                date_col = col
                break
        if date_col is None:
            return []
        # print(f"Date column index: {date_col}")

        # Check if the driver is available on the date
        available_drivers = []

        self.current_date = date

        for driver in drivers:
            print(driver.id)
            if driver.id in self.driver_id_map:
                row = self.driver_id_map[driver.id]
                cell_info = self.getCellInfo(row, date_col)
                # print(f'Before {driver.id} + {cell_info}')
                if cell_info["fill_color"] != self.unavailable_slot:
                    # self.writing(row, date_col, value="欠")
                    # update the last working time of the driver
                    self.updateLastWorkingTime(row, date_col, driver)
                    available_drivers.append(driver.id)
                # print(f'After {driver.id} + {cell_info}')

        self.wb.save(self.workbook_path)  # Save the workbook to persist changes
        print(f"Workbook saved to {self.workbook_path}")
        # return a list of available drivers ids
        return available_drivers
    
    def updateLastWorkingTime(self, row, date_col, driver):
        # get date from date_col
        date = self.getKeyFromValue(self.date_index_map, date_col)
        # find the previous date
        prev_date = date - timedelta(days=1)
        if prev_date not in self.date_index_map:
            driver.last_work_time = None
            return
        # print(f"Previous date: {prev_date}")
        # find the previous date column index
        prev_date_col = self.date_index_map[prev_date]
        # get the previous cell info
        prev_cell_info = self.getCellInfo(row, prev_date_col)
        # check if the previous cell is filled with red color
        if prev_cell_info["fill_color"] == self.unavailable_slot:
            # if so, update the last working time to None
            driver.last_work_time = None
        else:
            print(f'Have yesterday work record: {prev_cell_info}')
            if prev_cell_info["value"] == self.missing_slot:
                # raise Exception(f"Missing slot for {driver.id} on {prev_date}")
                driver.last_work_time = None
                return
            if prev_cell_info["value"] == "REF" or prev_cell_info["value"] == "VAC":
                driver.last_work_time = None
                return
            last_work_time = prev_cell_info["value"].split("-")[1]
            # Split the reading into hours and minutes
            if '.' in last_work_time:
                hours, minutes_fraction = str(last_work_time).split('.')
                hours = int(hours)
                minutes = int(float('0.' + minutes_fraction) * 60)
            else:
                hours = int(last_work_time)
                minutes = 0
            # Create a datetime object with the current date and the extracted time
            current_date = self.current_date
            result_datetime = datetime.combine(current_date, datetime.min.time()).replace(hour=hours, minute=minutes)
            driver.last_work_time = result_datetime

    def dayHasBeenProcessed(self, date: datetime):
        # Get the column index based on the date
        if date not in self.date_index_map:
            return False
        date_col = self.date_index_map[date]
        
        # Check it has been processed
        cell_info = self.getCellInfo(3, date_col)
        if cell_info["value"] == self.processed_slot:
            return True
        return False
    
    def convert_datetime_to_reading(self, dt):
        hours = dt.hour
        minutes = dt.minute
        
        if minutes == 0:
            return float(hours)
        else:
            return hours + minutes / 60.0

    def writeToMonthlyRoster(self, result:list, employees:dict, shifts:dict):
        # result = {s1.id: e1.id, s2: e4.id, s3.id: [e2.id, e3.id]}
        # Write the result to the monthly roster

        print("Employees:")
        print(employees)


        for shift_id, employee_id in result.items():
            # Get the shift and employee objects
            shift = shifts[shift_id]
            employee = employees[employee_id]

            # Get the row and column index based on the employee and shift
            row = self.driver_id_map[employee.id]
            col = self.date_index_map[self.current_date]

            print(f"Writing to cell ({row}, {col})")

            # Write the shift start and end time to the cell
            start_time = self.convert_datetime_to_reading(shift.start_time)
            end_time = self.convert_datetime_to_reading(shift.end_time)
            self.writing(row, col, value=f"{start_time}-{end_time}", fill_color="FFFFC000")

            # Mark the cell as processed
            self.writing(3, col, value=self.processed_slot, fill_color="FF00B050")

        self.wb.save(self.workbook_path)  # Save the workbook to persist changes






