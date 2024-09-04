from openpyxl import load_workbook
from openpyxl.styles import PatternFill
from openpyxl.utils import get_column_letter
from datetime import datetime, timedelta
from constants import OUTPUT_PATH
import os
import shutil

class DailyParser():
    
    # write a specific format of cell into the excel file
    # all cell in the given range if not filled with red color will be filled with that specific format
    def writing(self, row, col, value=None, fill_color="00000000"):
        cell = self.ws.cell(row=row, column=col)
        
        # Write a value to the cell if provided
        # color of the value will be black
        if value is not None:
            cell.value = value
            cell.font = cell.font.copy(color="FF000000")
        
        # Change the fill color of the cell
        cell.fill = PatternFill(start_color=fill_color, end_color=fill_color, fill_type="solid")
        print(f"Writing to cell ({row}, {col}): value={value}, fill_color={fill_color}")
    
        return cell


    def writeToNewDailySchedule(self, result:list, monthly_roster_dir:str, daily_schedule_filename:str, weekly_template_dir, employees:dict):
        # result = {s1.id: e1.id, s2: e4.id, s3.id: [e2.id, e3.id]}
        # Write the result to the daily schedule
        # check if monthly roaster dir exists
        if not os.path.exists(os.path.join(OUTPUT_PATH, monthly_roster_dir)):
            # if not exisit: create dir
            os.mkdir(os.path.join(OUTPUT_PATH, monthly_roster_dir))

        monthly_roster_dir = os.path.join(OUTPUT_PATH, monthly_roster_dir)
        # copy daily template for the day to dir named as daily_schedule _filename
        # overwrite if already exists
        daily_schedule_path = os.path.join(monthly_roster_dir, daily_schedule_filename)

        
        
        # template file path
        template_path = os.path.join(weekly_template_dir, daily_schedule_filename.split(" ")[0]+".xlsx")
        # print(f"Copying {template_path} to {daily_schedule_path}")

        shutil.copy(template_path, daily_schedule_path)


        # load the workbook
        self.wb = load_workbook(daily_schedule_path)
        self.ws = self.wb.active

        column_letter = get_column_letter(1)
        self.ws.column_dimensions[column_letter].width = 60 

        # match shift to employee
        for shift_id, employee_id in result.items():
            # if a list - write list to the cell and change cell background color to yellow
            result = ""
            # default grey
            fill_color = "00C0C0C0"
            if type(employee_id) == list:
                # Generate a list of employee names
                employee_names = [employees[id].name for id in employee_id]

                # Join the names with a comma
                result = ", ".join(employee_names)

                # Set the fill color to yellow
                fill_color = "FF6600"
            else :
                # get the name of the employee
                employee = employees[employee_id]
                result = employee.name
            self.writing(shift_id, 1, value=result, fill_color=fill_color)

            


        self.wb.save(daily_schedule_path)  # Save the workbook to persist changes