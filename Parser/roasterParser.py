from openpyxl import load_workbook
from openpyxl.styles import PatternFill
from datetime import datetime
import os
class RoasterParser():

    unavailable_slot = "FFFF0000"
    driver_id_map = {}
    date_index_map = {}
    ws = None

    def __init__(self, workbook_path):
        # self.driver_id_map = self.initDriverIdMap()
        # self.date_index_map = self.initDateIndexMap()
        # Load the workbook and select the active sheet
        self.workbook_path = os.path.join(os.getcwd(),"Input Data", workbook_path)
        self.wb = load_workbook(self.workbook_path)
        self.ws = self.wb.active
    


    # Function to access a cell based on row and column
    def getCellInfo(self, row, col):

        # Get the cell based on the row and column
        cell = self.ws.cell(row=row, column=col)
        cell_info = {
            "value": cell.value,
            "fill_color": None
        }

        # Check if the cell has a fill and access its color
        if cell.fill and cell.fill.fgColor.type == "rgb":
                cell_info["fill_color"] = cell.fill.fgColor.rgb

        return cell_info
    
    # write a specific format of cell into the excel file
    # all cell in the given range if not filled with red color will be filled with that specific format
    def writing(self, row, col, value=None):
        cell = self.ws.cell(row=row, column=col)
        
        # Write a value to the cell if provided
        # color of the value will be black
        if value is not None:
            cell.value = value
            cell.font = cell.font.copy(color="FF000000")
        
        # Change the fill color of the cell
        cell.fill = PatternFill(start_color="FFFFC000", end_color="FFFFC000", fill_type="solid")
        print(f"Writing to cell ({row}, {col}): value={value}, fill_color=FFFFC000")

        cell = self.ws.cell(row=row+1, column=col)

        # fill with empty color and empty value
        cell.fill = PatternFill(start_color="FFFFFFFF", end_color="FFFFFFFF", fill_type="solid")
        cell.value = None
    
        return cell

    # Precondition: 
    # - date is a datetime object
    # - drivers is a list of driver class objects
    # Postcondition:
    # - Returns a list of available drivers based on the date
    def getAvailableDrivers(self, date:datetime, drivers:list):
        # Get the column index based on the date
        date_col = None
        for col in range(1, self.ws.max_column + 1):
            print(self.ws.cell(row=2, column=col).value)
            if self.ws.cell(row=2, column=col).value == date:
                date_col = col
                break
        if date_col is None:
            return []
        # print(f"Date column index: {date_col}")
        
        # Get the row index mapping to the driver id
        driver_id_map = {}
        for row in range(2, self.ws.max_row + 1):
            driver_id = self.ws.cell(row=row, column=1).value
            if driver_id is not None:
                driver_id_map[self.ws.cell(row=row, column=1).value] = row
        print(driver_id_map)

        # Check if the driver is available on the date
        available_drivers = []

        for driver in drivers:
            print(driver.id)
            if driver.id in driver_id_map:
                row = driver_id_map[driver.id]
                cell_info = self.getCellInfo(row, date_col)
                # print(f'Before {driver.id} + {cell_info}')
                if cell_info["fill_color"] != self.unavailable_slot:
                    self.writing(row, date_col, value="欠")
                    available_drivers.append(driver.id)

                # print(f'After {driver.id} + {cell_info}')

        self.wb.save(self.workbook_path)  # Save the workbook to persist changes
        print(f"Workbook saved to {self.workbook_path}")
        # return a list of available drivers ids
        return available_drivers
    
    

if __name__ == "__main__":
     # Example usage
    parser = RoasterParser()
    # date = datetime(2023, 10, 1)  # Example date
    # drivers = []  # Example list of driver objects
    # available_drivers = parser.getAvailableDrivers(date, drivers)
    # print(available_drivers)

    # Example of writing to a cell
    workbook_path = os.path.join(os.getcwd(), "Input Data/roaster_M.xlsx")
    wb = load_workbook(workbook_path)
    ws = wb.active
    parser.writing(4, 4, ws, value="Test Value")
    wb.save(workbook_path)  # Save the workbook to persist changes
