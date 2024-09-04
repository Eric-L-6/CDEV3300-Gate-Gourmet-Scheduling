import tkinter as tk
from tkinter import ttk, messagebox
from openpyxl import load_workbook
import os
from datetime import datetime, timedelta

def update_name_dropdown():
    """Populate the name dropdown with non-empty names from the selected file."""
    selected_file = file_combobox.get()
    if not selected_file:
        messagebox.showwarning("No File Selected", "Please select a file first.")
        return
    
    try:
        wb = load_workbook(selected_file)
        ws = wb.active
        
        # Extract non-empty names from B9:B23
        names = []
        for cell in ws['B9:B23']:
            if cell[0].value:
                names.append(cell[0].value)
        
        # Update the name_combobox with the extracted names
        name_combobox['values'] = names
        if names:
            name_combobox.current(0)  # Select the first name by default
        else:
            name_combobox.set('')  # Clear the combobox if no names are found

    except Exception as e:
        messagebox.showerror("Error", f"An error occurred: {e}")

def calculate_days(ws):
    """Calculate the days based on the starting day in AC3 and populate the D6:AH6 range."""
    start_date = ws['AC3'].value
    if start_date is None:
        messagebox.showerror("Invalid Start Date", "AC3 does not contain a valid date.")
        return []

    if not isinstance(start_date, datetime):
        messagebox.showerror("Invalid Date Format", "AC3 does not contain a valid datetime object.")
        return []

    days = []
    current_date = start_date

    # Populate the days from D6 to AH6
    for col in range(4, 34):  # D is column 4, AH is column 34
        if current_date.day <= 31:
            days.append((ws.cell(row=6, column=col).column, str(current_date.day)))
            current_date += timedelta(days=1)
        else:
            break
    
    return days

def submit():
    selected_file = file_combobox.get()
    name = name_combobox.get()
    day = day_combobox.get()
    month = month_combobox.get()
    year = year_combobox.get()
    
    # Determine the status based on checkboxes
    if compensation_var.get():
        status = "off"
    elif refresh_var.get():
        status = "R"
    elif vacation_var.get():
        status = "V"
    elif normal_var.get():
        status = "OFF"
    else:
        status = ""

    if not day or not month or not year:
        messagebox.showwarning("Incomplete Input", "Please select a complete date.")
        return

    # Load the selected Excel workbook
    try:
        wb = load_workbook(selected_file)
        ws = wb.active
        
        # Find the row index for the name (B9:B23)
        name_row = None
        for cell in ws['B9:B23']:
            if cell[0].value == name:
                name_row = cell[0].row
                break
        
        if name_row is None:
            messagebox.showerror("Name Not Found", f"Name '{name}' not found in the range B9:B23.")
            return

        # Calculate the days and find the column index for the selected day
        days = calculate_days(ws)
        day_col = None
        for col, day_value in days:
            if day_value == day:
                day_col = col
                break
        
        if day_col is None:
            messagebox.showerror("Day Not Found", f"Day '{day}' not found in the calculated days range.")
            return

        # Update the cell with the status
        ws.cell(row=name_row, column=day_col, value=status)

        # Save the workbook
        wb.save(selected_file)
        
        # Show a confirmation message
        messagebox.showinfo("Success", f"Updated {name} for day {day} to '{status}' in {selected_file}")

        # Clear all fields
        name_combobox.set('')
        day_combobox.set('')
        month_combobox.set('')
        year_combobox.set('')
        compensation_var.set(False)
        refresh_var.set(False)
        vacation_var.set(False)
        normal_var.set(False)

    except FileNotFoundError:
        messagebox.showerror("File Not Found", f"The file '{selected_file}' was not found.")
    except Exception as e:
        messagebox.showerror("Error", f"An error occurred: {e}")

def get_excel_files():
    """Returns a list of .xlsx files in the current directory."""
    current_dir = os.getcwd()
    return [f for f in os.listdir(current_dir) if f.endswith('.xlsx')]

# Create the main window
root = tk.Tk()
root.title("Work Scheduler")

# Create and place widgets with grid layout
tk.Label(root, text="Select File:").grid(row=0, column=0, padx=10, pady=10, sticky='w')
file_combobox = ttk.Combobox(root, values=get_excel_files())
file_combobox.grid(row=0, column=1, padx=10, pady=10, sticky='ew')
file_combobox.bind("<<ComboboxSelected>>", lambda e: update_name_dropdown())  # Update names when a file is selected

tk.Label(root, text="Name:").grid(row=1, column=0, padx=10, pady=10, sticky='w')
name_combobox = ttk.Combobox(root)
name_combobox.grid(row=1, column=1, padx=10, pady=10, sticky='ew')

tk.Label(root, text="Day:").grid(row=2, column=0, padx=10, pady=10, sticky='w')
day_combobox = ttk.Combobox(root, values=[str(d) for d in range(1, 32)])
day_combobox.grid(row=2, column=1, padx=10, pady=10, sticky='ew')

tk.Label(root, text="Month:").grid(row=3, column=0, padx=10, pady=10, sticky='w')
month_combobox = ttk.Combobox(root, values=[f"{m:02d}" for m in range(1, 13)])
month_combobox.grid(row=3, column=1, padx=10, pady=10, sticky='ew')

tk.Label(root, text="Year:").grid(row=4, column=0, padx=10, pady=10, sticky='w')
year_combobox = ttk.Combobox(root, values=[str(y) for y in range(2020, 2031)])
year_combobox.grid(row=4, column=1, padx=10, pady=10, sticky='ew')

# Checkboxes for compensation, refresh, vacation, and normal statuses
compensation_var = tk.BooleanVar()
refresh_var = tk.BooleanVar()
vacation_var = tk.BooleanVar()
normal_var = tk.BooleanVar()

tk.Checkbutton(root, text="Compensation", variable=compensation_var).grid(row=5, column=0, padx=10, pady=10, sticky='w')
tk.Checkbutton(root, text="Refresh", variable=refresh_var).grid(row=5, column=1, padx=10, pady=10, sticky='w')
tk.Checkbutton(root, text="Vacation", variable=vacation_var).grid(row=6, column=0, padx=10, pady=10, sticky='w')
tk.Checkbutton(root, text="Normal", variable=normal_var).grid(row=6, column=1, padx=10, pady=10, sticky='w')

submit_button = tk.Button(root, text="Submit", command=submit)
submit_button.grid(row=7, column=0, columnspan=2, padx=10, pady=10)

# Configure column and row weights for resizing
root.grid_columnconfigure(1, weight=1)  # Allow column 1 to expand
root.grid_rowconfigure(0, weight=0)  # Fixed row height for row 0
root.grid_rowconfigure(1, weight=1)  # Allow row 1 to expand
root.grid_rowconfigure(2, weight=1)  # Allow row 2 to expand
root.grid_rowconfigure(3, weight=1)  # Allow row 3 to expand
root.grid_rowconfigure(4, weight=1)  # Allow row 4 to expand
root.grid_rowconfigure(5, weight=1)  # Allow row 5 to expand
root.grid_rowconfigure(6, weight=1)  # Allow row 6 to expand
root.grid_rowconfigure(7, weight=0)  # Fixed row height for row 7

# Run the application
root.mainloop()