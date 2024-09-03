import os
import openpyxl
from datetime import datetime

class Shift:
    def __init__(self, id, start_time, end_time, task_requirement):
        self.id = id
        self.start_time = start_time
        self.end_time = end_time
        self.task_requirement = task_requirement
    
    def __str__(self):
        return f"Shift: {self.id}\nStart time: {self.start_time}\nEnd time: {self.end_time}\nTask requirement: {self.task_requirement}\n"
    
class ShiftMaker:
    def shiftInitialise(filename):
        # load the workbook and select the active sheet
        workbook_path = os.path.join(os.getcwd(), "Input Data", filename)
        wb = openpyxl.load_workbook(workbook_path)
        ws = wb.active
        shifts = []

        # Start from the column of Task requirement
        task_requirement_col = 1
        while ws.cell(row=1, column=task_requirement_col).value != "Task Requirement":
            task_requirement_col += 1
        
        index = 2
        while ws.cell(index, task_requirement_col).value != None:
            task_requirement = ws.cell(index, task_requirement_col).value.split(";")
            time_range = ws.cell(index, task_requirement_col+1).value.split("-")
            start_time_str = time_range[0].strip()
            end_time_str = time_range[1].strip()

            # Convert time strings to datetime objects
            start_time = datetime.strptime(start_time_str, "%H:%M")
            end_time = datetime.strptime(end_time_str, "%H:%M")

            new_shift = Shift(index, start_time, end_time, task_requirement)
            shifts.append(new_shift)
            index += 2

        return shifts


if __name__ == "__main__":
    shifts = ShiftMaker.shiftInitialise("daily_M.xlsx")
    for shift in shifts:
        print(shift)
